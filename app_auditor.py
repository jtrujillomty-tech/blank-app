import streamlit as st
import pdfplumber
import pandas as pd
import numpy as np
import re
import io
import plotly.express as px
import logging
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================
# 🛑 SILENCIADOR DE TERMINAL
# ==========================================
logging.getLogger("pdfminer").setLevel(logging.WARNING)

# ==========================================
# CONFIGURACIÓN DE LA PÁGINA Y FIRMA
# ==========================================
st.set_page_config(page_title="Hub de Auditoría", page_icon="🕵️‍♂️", layout="wide")

st.sidebar.markdown(
    """
    <div style="position: fixed; bottom: 10px; left: 10px; font-size: 10px; color: gray;">
        V3.5 - Desarrollado por JJ para auditorias de facturacion Aduax
    </div>
    """, 
    unsafe_allow_html=True
)

# ==========================================
# 🧠 FUNCIONES: MÓDULO DE FACTURACIÓN
# ==========================================
def extraer_tablas_pdf(pdf_file, texto_estado, barra_progreso):
    todas_las_tablas = []
    with pdfplumber.open(pdf_file) as pdf:
        total_paginas = len(pdf.pages)
        for i, pagina in enumerate(pdf.pages):
            texto_estado.markdown(f"**📄 Escaneando PDF:** Página {i+1} de {total_paginas}...")
            barra_progreso.progress((i) / total_paginas)
            
            tablas_pagina = pagina.extract_tables()
            for tabla in tablas_pagina:
                df_tabla = pd.DataFrame(tabla).replace('\n', ' ', regex=True)
                todas_las_tablas.append(df_tabla)
                
    barra_progreso.progress(1.0)
    if todas_las_tablas:
        return pd.concat(todas_las_tablas, ignore_index=True)
    return pd.DataFrame()

def estructurar_tabulador(df_raw, texto_estado):
    texto_estado.markdown("**🧠 Estructurando tarifas, observaciones y claves...**")
    datos_limpios = []
    impoexp_actual = None
    aduasecc_actual = None
    claves_actual = "TODAS"
    
    for index, row in df_raw.iterrows():
        fila_texto = " | ".join([str(val).strip() for val in row if pd.notna(val)])
        fila_texto_upper = fila_texto.upper()
        
        if 'IMPORTACI' in fila_texto_upper or 'EXPORTACI' in fila_texto_upper or 'ESPECIALES' in fila_texto_upper:
            if 'IMPORTACI' in fila_texto_upper: impoexp_actual = 'Importación'
            elif 'EXPORTACI' in fila_texto_upper: impoexp_actual = 'Exportación'
            else: impoexp_actual = 'Especiales'
            
            numeros_aduana = re.findall(r'\b\d{3}\b', fila_texto)
            numeros_aduana = [num for num in numeros_aduana if num != '999']
            
            if 'TODAS' in fila_texto_upper and not numeros_aduana:
                aduasecc_actual = "TODAS"
            else:
                aduasecc_actual = ", ".join(numeros_aduana) if numeros_aduana else "GENERAL"
            
            claves_encontradas = re.findall(r'\b[A-Z0-9]{2}\b|\b000\b|\b999\b|-999', fila_texto_upper)
            stop_words = ['EN', 'DE', 'LA', 'EL', 'PO', 'SI', 'NO', 'MX', 'US', 'AL', 'SE', 'SU']
            claves_validas = [c for c in claves_encontradas if (c not in stop_words and not c.isdigit()) or c in ['000', '999', '-999']]
            
            if any(c in ['000', '999', '-999'] for c in claves_validas):
                claves_actual = "TODAS"
            else:
                claves_actual = ", ".join(claves_validas) if claves_validas else "TODAS"
            continue
        
        if impoexp_actual and 'PREVIO EN ORIGEN' in fila_texto_upper:
            numeros_obs = re.findall(r'[\$]?\s*(\d+([.,]\d+)?)', fila_texto_upper)
            if numeros_obs:
                tarifa_obs = float(numeros_obs[0][0])
                if tarifa_obs > 0:
                    datos_limpios.append({
                        'IMPOEXP': impoexp_actual, 'ADUASECC': aduasecc_actual, 'CLAVES': claves_actual,
                        'CONCEPTO': 'OPECTE104 PREVIO EN ORIGEN', 'TARIFA': tarifa_obs, 'FACTOR': 'OPERACION'
                    })
        
        if impoexp_actual and ('OPECTE' in fila_texto_upper or 'HONORARIOS' in fila_texto_upper):
            concepto = ""; tarifa = 0.0; factor = ""
            for val in row:
                val_str = str(val).strip() if pd.notna(val) else ""
                val_upper = val_str.upper()
                if 'OPECTE' in val_upper or 'HONORARIOS' in val_upper: concepto = val_str
                elif any(palabra in val_upper for palabra in ['CONTENEDOR', 'PEDIMENTO', 'LINEA', 'CAJA', 'ADICIONAL', 'FCL', 'LCL', 'CARGA']): factor = val_str
                elif re.fullmatch(r'[\$]?\d+([.,]\d+)?', val_str):
                    tarifa_limpia = re.sub(r'[^\d.]', '', val_str)
                    if tarifa_limpia and tarifa == 0.0: tarifa = float(tarifa_limpia)

            if tarifa > 0:
                datos_limpios.append({
                    'IMPOEXP': impoexp_actual, 'ADUASECC': aduasecc_actual, 'CLAVES': claves_actual,
                    'CONCEPTO': concepto, 'TARIFA': tarifa, 'FACTOR': factor
                })
                
    return pd.DataFrame(datos_limpios)

def auditar_facturacion(df_ventas, df_tabulador, df_os, texto_estado, barra_progreso):
    df_tabulador = df_tabulador[df_tabulador['CONCEPTO'].str.contains('OPECTE', na=False, case=False)].copy()
    df_tabulador['IMPOEXP_NORM'] = df_tabulador['IMPOEXP'].astype(str).str.upper().str.replace('Ó', 'O')
    resultados = []
    
    total_filas = len(df_ventas)
    claves_especiales = ['V1', 'V5', 'A3', 'A4', 'F4', 'F5', 'R1', 'GC']
    
    for index, operacion in df_ventas.iterrows():
        if index % 10 == 0: 
            texto_estado.markdown(f"**⚙️ Auditando Operaciones:** Revisando fila {index+1} de {total_filas}...")
            barra_progreso.progress(index / total_filas)
            
        pedimento = str(operacion.get('PEDNMBR', 'N/A')).strip()
        clave_ped = str(operacion.get('CLPEID', '')).strip().upper() 
        referencia = str(operacion.get('OPREFEREN', 'N/A'))
        cliente_id = str(operacion.get('CUSTNMBR', ''))
        aduana_op = str(operacion.get('ADUASECC', '')).replace('.0', '')
        impoexp_op = str(operacion.get('IMPOEXP', '')).strip().upper()
        item_op = str(operacion.get('ITEMNMBR', ''))
        descripcion_op = str(operacion.get('GRPDESCE', 'N/A')) 
        formula = str(operacion.get('FORMULA', ''))
        
        aux_factura = str(operacion.get('AUX1', ''))
        prenumbe = str(operacion.get('PRENUMBE', ''))
        texto_busqueda_empresa = f"{referencia} {aux_factura} {prenumbe}".upper()
        empresa = "EILO" if "ANLD" in texto_busqueda_empresa else "ELI"
        
        monto_cobrado = float(operacion.get('OXTNDPRC', 0.0))
        if monto_cobrado == 0: continue
        
        moneda = str(operacion.get('CURNCYID', '')).strip().upper()
        tc = float(operacion.get('XCHGRATE', 1.0))
        if tc <= 0: tc = 1.0
        
        cant_it = {
            'CONTENEDOR': float(operacion.get('contenedores', 0.0)),
            'LINEA': float(operacion.get('lineas de facturas', 0.0)),
            'GUIA MASTER': float(operacion.get('Guias_M', 0.0)),
            'GUIA HOUSE': float(operacion.get('Guias_H', 0.0))
        }
        
        fila_reporte = {
            'Referencia': referencia, 'Pedimento': pedimento, 'Aduana': aduana_op, 
            'Clave': clave_ped, 'Operacion': impoexp_op, 'Articulo': item_op, 
            'Descripcion': descripcion_op, 'Moneda': moneda, 'TC': tc, 
            'Monto_Cobrado': monto_cobrado, 'Cantidad_Fisica_IT': 'No Aplica', 
            'Categoria': '', 'Estatus': '', 'Detalle': '', 'Empresa': empresa
        }

        # REGLAS FAST-TRACK
        if item_op == 'OPECTE070':
            fila_reporte.update({'Categoria': 'OK, todo en orden', 'Estatus': '✅ Correcto', 'Detalle': 'No esta en tabulador (Cobro default válido).'})
            resultados.append(fila_reporte)
            continue
        if item_op == 'OPECTE047':
            fila_reporte.update({'Categoria': 'OK, todo en orden', 'Estatus': '✅ Correcto', 'Detalle': 'Servicios de Compliance.'})
            resultados.append(fila_reporte)
            continue
        if item_op in ['OPECTE021', 'OPECTE061']:
            sin_pedimento = (not pedimento or pedimento == 'NAN' or pedimento == 'N/A' or pedimento == '')
            sin_clave = (not clave_ped or clave_ped == 'NAN' or clave_ped == '')
            if sin_pedimento and sin_clave:
                fila_reporte.update({'Categoria': 'OK, todo en orden', 'Estatus': '✅ Correcto', 'Detalle': 'Operación sin pedimento ni clave (Validado por regla de negocio).'})
                resultados.append(fila_reporte)
                continue

        tipo_op_buscar = 'IMPORTACION' if impoexp_op.startswith('IMP') else 'EXPORTACION' if impoexp_op.startswith('EXP') else impoexp_op

        if not clave_ped or clave_ped == 'NAN' or clave_ped == '':
            fila_reporte.update({'Categoria': 'Concepto no tabulado', 'Estatus': '🚨 Error', 'Detalle': "🚨 Clave de pedimento en blanco en el reporte."})
            resultados.append(fila_reporte)
            continue

        if 'SELECT' in formula.upper():
            match_select = re.search(r'SELECT\s+(\d+([.,]\d+)?)', formula.upper())
            tarifa_select = float(match_select.group(1)) if match_select else 0.0
            encontrado_os = False
            if not df_os.empty and tarifa_select > 0:
                filtro_os = df_os.astype(str).apply(lambda row: str(cliente_id) in row.values and str(tarifa_select) in row.values, axis=1)
                if filtro_os.any(): encontrado_os = True

            if encontrado_os and abs(monto_cobrado - tarifa_select) <= 0.01:
                fila_reporte.update({'Categoria': 'OK, todo en orden', 'Estatus': '✅ Correcto', 'Detalle': "SELECT validado vs Catálogo OS."})
            else:
                fila_reporte.update({'Categoria': 'Revisión Manual OS', 'Estatus': '⚠️ Alerta', 'Detalle': f"Fórmula SELECT con ${tarifa_select}. Validar manualmente."})
            resultados.append(fila_reporte)
            continue

        # EXTRACCIÓN Y PUNTUACIÓN DE CANDIDATOS
        mask_impo = (df_tabulador['IMPOEXP_NORM'] == tipo_op_buscar)
        mask_esp = (df_tabulador['IMPOEXP_NORM'] == 'ESPECIALES') if clave_ped in claves_especiales else False
        mask_r1_cross = (clave_ped == 'R1') & (df_tabulador['CLAVES'].str.contains('R1', na=False) | df_tabulador['CONCEPTO'].str.contains('RECTIFICACION', case=False, na=False))
        mask_aduanas = df_tabulador['ADUASECC'].astype(str).str.contains(aduana_op, na=False) | df_tabulador['ADUASECC'].astype(str).str.contains(r'\b000\b|TODAS|GENERAL', na=False, case=False)
        
        if item_op == 'OPECTE050':
            mask_item = df_tabulador['CONCEPTO'].str.contains('OPECTE050', na=False, case=False) | df_tabulador['CONCEPTO'].str.contains('ADICIONAL', na=False, case=False) | df_tabulador['FACTOR'].str.contains('ADICIONAL', na=False, case=False)
        else:
            mask_item = df_tabulador['CONCEPTO'].str.contains(item_op, na=False, case=False)
        
        def check_clave(val):
            if not isinstance(val, str) or val == "TODAS" or "000" in val: return True
            return clave_ped in val

        mask_clave = df_tabulador['CLAVES'].apply(check_clave)
        
        candidatos = df_tabulador[((mask_impo | mask_esp | mask_r1_cross) & mask_aduanas & mask_item & mask_clave)].copy()
        
        if candidatos.empty:
            fila_reporte.update({'Categoria': 'Concepto no tabulado', 'Estatus': '🚨 Error', 'Detalle': f"No se encontró tarifa en el PDF (Validado Operación y Clave {clave_ped})."})
            resultados.append(fila_reporte)
            continue

        candidatos['SCORE'] = 0
        for idx, row_c in candidatos.iterrows():
            score = 0
            if clave_ped in str(row_c['CLAVES']): score += 20 
            if aduana_op in str(row_c['ADUASECC']): score += 10 
            if clave_ped == 'R1' and 'RECTIFICACION' in str(row_c['CONCEPTO']).upper(): score += 10
            
            factor_pdf_c = str(row_c['FACTOR']).upper()
            if 'CONTENEDOR' in formula.upper() or 'FCL' in formula.upper():
                if 'CONTENEDOR' in factor_pdf_c or 'FCL' in factor_pdf_c: score += 5
            elif 'SUELTA' in formula.upper() or 'LCL' in formula.upper() or 'CARGA' in formula.upper():
                if 'SUELTA' in factor_pdf_c or 'LCL' in factor_pdf_c or 'CARGA' in factor_pdf_c: score += 5
            
            candidatos.at[idx, 'SCORE'] = score
            
        candidatos = candidatos.sort_values(by='SCORE', ascending=False)

        # LOOP MATEMÁTICO
        reporte_generado = False
        tarifas_intentadas = []
        
        for idx, row_tarifa in candidatos.iterrows():
            tarifa_base = float(row_tarifa['TARIFA'])
            factor_pdf = str(row_tarifa['FACTOR']).upper()
            tarifas_intentadas.append(str(tarifa_base))
            
            division_directa = monto_cobrado / tarifa_base if tarifa_base > 0 else 0
            es_entero_directo = abs(division_directa - round(division_directa)) <= 0.001
            
            monto_convertido = monto_cobrado / tc if tc > 1.5 else monto_cobrado
            division_tc = monto_convertido / tarifa_base if tarifa_base > 0 else 0
            es_entero_tc = abs(division_tc - round(division_tc)) <= 0.001
            
            if es_entero_directo or (es_entero_tc and 'MX' in moneda):
                division = division_directa if es_entero_directo else division_tc
                nota_moneda = f" (Aplicando TC) ({monto_convertido:,.2f} USD)" if (not es_entero_directo and es_entero_tc) else ""
                multiplo_facturado = round(division)
                valido_it = False
                
                if 'CAJA' in factor_pdf or 'CONTENEDOR' in factor_pdf:
                    cantidad_real = cant_it['CONTENEDOR']
                    fila_reporte['Cantidad_Fisica_IT'] = f"{cantidad_real} Unidad(es)"
                    valido_it = True
                    
                    if multiplo_facturado > cantidad_real: 
                        if ('LCL' in factor_pdf or 'CONSOLIDADA' in factor_pdf) and cantidad_real == 0:
                            fila_reporte.update({'Categoria': 'OK, todo en orden', 'Estatus': '✅ Correcto', 'Detalle': f"Match perfecto (Excepción LCL aplicada a tarifa ${tarifa_base}).{nota_moneda}"})
                        else:
                            fila_reporte.update({'Categoria': 'Se facturó de MÁS', 'Estatus': '🚨 Sobrecobro', 'Detalle': f"Múltiplo de tarifa ${tarifa_base}: cobraron {multiplo_facturado} pero IT reporta solo {cantidad_real}.{nota_moneda}"})
                    elif multiplo_facturado < cantidad_real: 
                        fila_reporte.update({'Categoria': 'Se facturó de MENOS', 'Estatus': '🚨 Subcobro', 'Detalle': f"Múltiplo de tarifa ${tarifa_base}: cobraron {multiplo_facturado} pero IT reporta {cantidad_real}.{nota_moneda}"})
                    else: 
                        fila_reporte.update({'Categoria': 'OK, todo en orden', 'Estatus': '✅ Correcto', 'Detalle': f"Match perfecto vs Contenedores (Tarifa ${tarifa_base}).{nota_moneda}"})
                else:
                    for clave_it, cantidad_real in cant_it.items():
                        if clave_it in factor_pdf and cantidad_real > 0:
                            fila_reporte['Cantidad_Fisica_IT'] = f"{cantidad_real} {clave_it}(s)"
                            valido_it = True
                            if multiplo_facturado > cantidad_real: fila_reporte.update({'Categoria': 'Se facturó de MÁS', 'Estatus': '🚨 Sobrecobro', 'Detalle': f"Cobraron {multiplo_facturado}x ${tarifa_base} pero IT reporta {cantidad_real}.{nota_moneda}"})
                            elif multiplo_facturado < cantidad_real: fila_reporte.update({'Categoria': 'Se facturó de MENOS', 'Estatus': '🚨 Subcobro', 'Detalle': f"Cobraron {multiplo_facturado}x ${tarifa_base} pero IT reporta {cantidad_real}.{nota_moneda}"})
                            else: fila_reporte.update({'Categoria': 'OK, todo en orden', 'Estatus': '✅ Correcto', 'Detalle': f"Match perfecto vs IT (Tarifa ${tarifa_base}).{nota_moneda}"})
                            break
                
                if not valido_it:
                    fila_reporte.update({'Categoria': 'OK, todo en orden', 'Estatus': '✅ Correcto', 'Detalle': f"Múltiplo exacto ({multiplo_facturado}x) basado en tarifa ${tarifa_base}.{nota_moneda}"})
                
                resultados.append(fila_reporte)
                reporte_generado = True
                break 
                
        if not reporte_generado:
            resumen_tarifas = " | $".join(list(dict.fromkeys(tarifas_intentadas))) 
            fila_reporte.update({'Categoria': 'Error Múltiplo Incoherente', 'Estatus': '🚨 Error', 'Detalle': f"El cobro no es múltiplo de NINGUNA tarifa tabulada encontrada (${resumen_tarifas})."})
            resultados.append(fila_reporte)

    barra_progreso.progress(1.0)
    texto_estado.markdown("**✅ Análisis completado. Generando archivos y gráficos...**")

    df_completo = pd.DataFrame(resultados)
    df_ok = df_completo[df_completo['Categoria'] == 'OK, todo en orden']
    df_errores = df_completo[df_completo['Categoria'] != 'OK, todo en orden']
    
    resumen = pd.DataFrame()
    if not df_completo.empty:
        resumen = df_completo.groupby('Categoria').agg(Operaciones=('Referencia', 'count'), Monto_Involucrado=('Monto_Cobrado', 'sum')).reset_index().sort_values(by='Operaciones', ascending=False)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        if not resumen.empty: resumen.to_excel(writer, sheet_name='📊 Resumen', index=False)
        if not df_errores.empty: df_errores.to_excel(writer, sheet_name='🚨 Hallazgos', index=False)
        if not df_ok.empty: df_ok.to_excel(writer, sheet_name='✅ Correctas', index=False)
    output.seek(0)
    
    return output, df_completo

# ==========================================
# 🧠 FUNCIONES: MÓDULO CRUCES EILO (TU SCRIPT ORIGINAL)
# ==========================================
CLIENTES_SIN_TRASLADO = [
    "0003-TECNO MAIZ SA DE CV", "10230-FIBERS MEXICO HOLDINGS S DE RL DE CV", "4976-CRANE WORLDWIDE LOGISTICS, LLC",
    "5038-DSV AIR & SEA INC", "5628-SCHAEFFLER GROUP USA INC", "5791-YAMAHA MOTOR MANUFACTURING CORPORATION",
    "5983-KUEHNE & NAGEL", "0010-CRIBAS Y EQUIPOS INDUSTRIALES SA DE CV", "0308-DANIEL B HASTINGS INC",
    "0462-KLOECKNER METALS SERVICE CENTERS DE MEXICO S DE RL DE CV", "10020-MOTORES ELECTRICOS SUMERGUBLES DE MEXICO S DE RL DE CV(SERVICIOS",
    "10044-CONTINENTAL RETREAD MORELIA SA DE CV", "10081-PARKER HANNIFIN CORPORATION", "10210-KLOECKNER METALS SERVICE CENTERS DE MEXICO S DE RL DE CV",
    "10217-MOTORES ELECTRICOS SUMERGIBLES DE MEXICO S DE RL DE CV", "10241-PENTAIR COMMERCIAL ICE LLC", "10268-FRANKLIN ELECTRIC CO. INC.",
    "10609-MANITOWOC FOODSERVICE (SWITZERLAND) GMBH", "10649-FRANKLIN FUELING SYSTEMS LLC", "10731-SCHAEFFLER TRANSMISSION SYSTEMS, LLC",
    "10977-VECTORIAL NETWORK LOGISTICS LLC", "11042-KALPA LOGISTICS SA DE CV", "11178-EI CUSTOMS BROKER LLC",
    "11793-AL-COMINTERNATIONAL TRADE, INC.", "11830-DAGA WAREHOUSE SERVICE CORP", "11837-SCHAEFFLER SPECIAL MACHINERY LLC",
    "11845-WDM WATER SYSTEMS SA DE CV", "11888-SHIPX MEXICO SA DE CV", "11893-IEES INC", "4981-RLR INDUSTRIES INC.",
    "4987-LEONI WIRING SYSTEMS INC.", "4988-CLAYTON METALS INC.", "4992-ACUITY BRANDS LIGHTING INC.",
    "5029-INTERNATIONAL CUSTOMHOUSE BROKERAGE SERVICES  LTD", "5032-TOC LOGISTICS INTERNATIONAL INC", "5218-FREIG CARRILLO FORWARDING INC.",
    "5272-FORNEY CORPORATION", "5292-GLOBAL FORWARDING INC", "5662-CANTERA DOORS INC", "5796-ECONOMIZER MONTERREY SA DE CV",
    "6102-ARNECOM SA DE CV", "6789-AUTOMOTIVE VERITAS DE MEXICO SA DE CV", "6922-MARY KAY COSMETICS DE MEXICO SA DE CV",
    "7168-PEXCO OPERACIONES S DE RL DE CV", "7196-CONTINENTAL TIRE DE MEXICO S DE RL DE CV", "7293-IMPORTADORA BBBMEX S DE RL DE CV",
    "7514-MOTORES FRANKLIN, S.A. DE C.V.", "7793-DURAKON INDUSTRIES MEXICO SA DE CV", "7956-SCHAEFFLER TRANSMISION S DE RL DE CV",
    "9057-MOLINOS AZTECA SA DE CV", "9375-ACUITY BRANDS LIGHTING DE MEXICO S DE RL DE CV", "9410-HOLOPHANE SA DE CV",
    "9930-MOTORES ELECTRICOS SUMERGIBLES DE MEXICO S DE RL DE CV", "9934-MOTORES ELECTRICOS SUMERGIBLES DE MEXICO S DE RL DE CV (GPE)",
    "9853-ROCKWELL AUTOMATION DE MEXICO SA DE CV", "9755-MOLINOS AZTECA DE CHIAPAS SA DE CV", "9756-MOLINOS AZTECA DE VERACRUZ SA DE CV",
    "9761-HARINERA DE VERACRUZ SA DE CV", "9762-HARINERA DE MAIZ DE MEXICALI SA DE CV", "9777-MOLINOS AZTECA DE CHALCO SA DE CV",
    "9855-TRANSPORTE AEREO TECNICO EJECUTIVO SA DE CV", "11060-MOTION MEXICO S DE RL DE CV"
]

def limpiar_folio(folio):
    if pd.isna(folio): return ""
    numeros = re.findall(r'\d+', str(folio))
    if numeros: return str(int("".join(numeros)))
    return ""

def cargar_reporte(file_uploader):
    try:
        return pd.read_excel(file_uploader, skiprows=1)
    except Exception:
        file_uploader.seek(0)
        try:
            return pd.read_csv(file_uploader, encoding='latin1', skiprows=1, sep=',')
        except Exception:
            file_uploader.seek(0)
            return pd.read_csv(file_uploader, encoding='latin1', sep=',')

def auditar_cruces_eilo(file_facturas, file_compras, barra_progreso):
    df_facturas = cargar_reporte(file_facturas)
    df_compras = cargar_reporte(file_compras)

    df_facturas['Num_Base'] = df_facturas['Factura'].apply(limpiar_folio)
    df_compras['Num_Factura_Limpio'] = df_compras['Factura'].apply(limpiar_folio)

    if 'OrdenCompra' in df_compras.columns:
        df_compras['Num_OC_Limpio'] = df_compras['OrdenCompra'].apply(limpiar_folio)
    else:
        df_compras['Num_OC_Limpio'] = ""

    resultados_estatus = []
    resultados_comentarios = []
    total_filas = len(df_facturas)

    for index, row in df_facturas.iterrows():
        if total_filas > 0 and index % max(1, total_filas//10) == 0:
            barra_progreso.progress(index / total_filas)

        num_base = row['Num_Base']
        cliente_actual = row['Cliente'] if 'Cliente' in df_facturas.columns else ""
        referencia = row['Referencia'] if 'Referencia' in df_facturas.columns else None
        total_facturado = float(row['Total']) if pd.notna(row['Total']) else 0.0

        estatus_eilo = str(row.get('Estatus', '')).strip().lower()
        es_eilo_cancelada = 'cancel' in estatus_eilo
        es_eilo_activa = not es_eilo_cancelada

        if cliente_actual in CLIENTES_SIN_TRASLADO:
            if es_eilo_cancelada:
                resultados_estatus.append("Cancelada en EILO (Cliente americano)")
                resultados_comentarios.append("Exceptuado por lista. Registro cancelado en origen.")
            else:
                resultados_estatus.append("Cliente americano, no se requiere traslado")
                resultados_comentarios.append("Exceptuado automáticamente por lista de exclusión.")
            continue 

        match_compras = df_compras[
            (df_compras['Num_Factura_Limpio'] == num_base) | 
            ((df_compras['Num_OC_Limpio'] == num_base) & (df_compras['Num_OC_Limpio'] != ""))
        ]

        if match_compras.empty and pd.notna(referencia) and 'Referencia' in df_compras.columns:
            posibles_matches = df_compras[df_compras['Referencia'] == referencia]
            if not posibles_matches.empty and 'Total' in posibles_matches.columns:
                try:
                    posibles_matches['Total_Float'] = pd.to_numeric(posibles_matches['Total'], errors='coerce').fillna(0)
                    match_con_monto = posibles_matches[abs(posibles_matches['Total_Float'] - total_facturado) <= 1.0]
                    if not match_con_monto.empty:
                        match_compras = match_con_monto 
                except:
                    pass

        if match_compras.empty:
            if es_eilo_cancelada:
                resultados_estatus.append("Cancelada en EILO y no registrada en ELI")
                resultados_comentarios.append("OK. Sin impacto.")
            else:
                resultados_estatus.append("No encontrada en ELI")
                resultados_comentarios.append("Posible traslado no requerido.")
        else:
            comentarios = []
            estatus_final = "Encontrada"

            estatus_unicos_eli = match_compras['Estatus'].dropna().unique() if 'Estatus' in match_compras.columns else []
            estatus_eli_lower = [str(x).strip().lower() for x in estatus_unicos_eli]

            es_eli_cancelada = any('cancel' in x for x in estatus_eli_lower)
            es_eli_activa = any(x in ['abierto', 'contabilizada'] for x in estatus_eli_lower)
            es_eli_totalmente_cancelada = es_eli_cancelada and not es_eli_activa

            if es_eilo_activa and es_eli_totalmente_cancelada:
                estatus_final = "Alerta: Activa en EILO pero cancelada en ELI"
                comentarios.append("Desfasaje: Factura activa en EILO pero registro cancelado en ELI.")
            elif es_eilo_cancelada and es_eli_activa:
                estatus_final = "Alerta: Cancelada en EILO y activa en ELI"
                comentarios.append("Desfasaje: Factura americana cancelada en EILO y registro activo en ELI.")
            elif es_eilo_cancelada and es_eli_totalmente_cancelada:
                hay_refacturacion = False
                if pd.notna(referencia) and 'Referencia' in df_facturas.columns:
                    mismas_ref = df_facturas[df_facturas['Referencia'] == referencia]
                    activas_ref = mismas_ref[mismas_ref['Estatus'].astype(str).str.lower().str.contains('activ', na=False)]
                    if not activas_ref.empty:
                        hay_refacturacion = True
                if hay_refacturacion:
                    estatus_final = "OK cancelada en ambos lugares pero con refacturacion activa"
                else:
                    estatus_final = "OK cancelada en ambos lugares"
            else:
                if 'Moneda' in match_compras.columns:
                    if not (match_compras['Moneda'] == 'USD').all():
                        estatus_final = "ERROR CRÍTICO"
                        comentarios.append("La compra NO está en USD.")

                if 'Total' in match_compras.columns:
                    try:
                        totales = pd.to_numeric(match_compras['Total'], errors='coerce').fillna(0)
                        total_compras = totales.max()
                        if abs(total_facturado - total_compras) > 1.0:
                            comentarios.append(f"Diferencia en totales: Facturado {total_facturado} vs Compra {total_compras}.")
                    except: pass

                tiene_factura = match_compras['Factura'].notna().any() if 'Factura' in match_compras.columns else False
                tiene_oc = match_compras['OrdenCompra'].notna().any() if 'OrdenCompra' in match_compras.columns else False

                if estatus_final != "ERROR CRÍTICO":
                    if tiene_factura and 'contabilizada' in estatus_eli_lower:
                        estatus_final = "Contabilizada OK"
                    elif tiene_oc and ('abierto' in estatus_eli_lower or 'contabilizada' not in estatus_eli_lower):
                        po_detectadas = match_compras['OrdenCompra'].dropna().unique()
                        po_str = ", ".join([str(p) for p in po_detectadas])
                        estatus_final = f"OC Generada pero falta contabilizar ({po_str})"
                        comentarios.append("Falta folio de factura ANLD en Columna B.")

            resultados_estatus.append(estatus_final)
            resultados_comentarios.append(" | ".join(comentarios))

    df_facturas['Auditoria_Estatus'] = resultados_estatus
    df_facturas['Auditoria_Comentarios'] = resultados_comentarios

    if 'Num_Base' in df_facturas.columns: df_facturas = df_facturas.drop(columns=['Num_Base'])

    df_resumen = df_facturas['Auditoria_Estatus'].value_counts().reset_index()
    df_resumen.columns = ['Estatus de Auditoría', 'Cantidad de Eventos']

    columnas_interes = ['Cliente', 'Factura', 'Referencia', 'Fecha', 'Moneda', 'Total', 'Auditoria_Estatus', 'Auditoria_Comentarios']
    columnas_existentes = [col for col in columnas_interes if col in df_facturas.columns]
    df_simplificada = df_facturas[columnas_existentes].drop_duplicates(subset=['Factura']).copy()

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_resumen.to_excel(writer, sheet_name='Resumen', index=False)
        df_simplificada.to_excel(writer, sheet_name='Simplificada', index=False)
        df_facturas.to_excel(writer, sheet_name='Detalle', index=False)
        
        azul_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        fuente_blanca_negrita = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        alineacion_centro = Alignment(horizontal="center", vertical="center")
        borde_delgado = Border(
            left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
        )

        for ws_name in ['Resumen', 'Simplificada', 'Detalle']:
            ws = writer.sheets[ws_name]
            ws.views.sheetView[0].showGridLines = True
            for cell in ws[1]:
                cell.fill = azul_header
                cell.font = fuente_blanca_negrita
                cell.alignment = alineacion_centro
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = borde_delgado
                    if ws_name == 'Resumen' and cell.column == 2:
                        cell.alignment = alineacion_centro
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    barra_progreso.progress(1.0)
    return output, df_resumen

# ==========================================
# 🖥️ INTERFAZ GRÁFICA PRINCIPAL (HUB)
# ==========================================
st.sidebar.title("🗂️ Módulos de Auditoría")
opcion_auditor = st.sidebar.radio(
    "Selecciona la herramienta a utilizar:",
    ["Auditor de Facturación", "Auditor Cruces EILO", "Auditor ELI (Próximamente)"]
)

if opcion_auditor == "Auditor de Facturación":
    st.title("🕵️‍♂️ Plataforma de auditoria Facturacion")
    st.markdown("Sube los archivos necesarios para realizar la auditoría y visualizar los resultados interactivos.")

    col1, col2 = st.columns(2)

    with col1:
        archivo_pdf = st.file_uploader("📄 1. PDF del Tabulador", type=["pdf"])
        archivo_os = st.file_uploader("📂 3. Catálogo OS (Opcional)", type=["xlsx", "xls"])

    with col2:
        archivo_excel = st.file_uploader("📊 2. Excel (Funcion Select)", type=["xlsx", "xls", "csv"])

    st.divider()

    if st.button("🚀 Ejecutar Auditoría de Facturación", use_container_width=True):
        if archivo_pdf and archivo_excel:
            espacio_texto = st.empty()
            barra_progreso = st.progress(0)
            
            try:
                df_ventas = pd.read_excel(archivo_excel)
                df_os = pd.read_excel(archivo_os) if archivo_os else pd.DataFrame()
                
                df_raw = extraer_tablas_pdf(archivo_pdf, espacio_texto, barra_progreso)
                df_tabulador = estructurar_tabulador(df_raw, espacio_texto)
                
                barra_progreso.progress(0)
                excel_reporte, df_completo = auditar_facturacion(df_ventas, df_tabulador, df_os, espacio_texto, barra_progreso)
                
                espacio_texto.empty()
                barra_progreso.empty()
                st.success("✅ ¡Auditoría de Facturación completada exitosamente!")
                
                st.markdown("---")
                st.header("📊 Dashboard de Resultados")
                
                if not df_completo.empty:
                    kpi1, kpi2, kpi3 = st.columns(3)
                    total_operaciones = len(df_completo)
                    total_facturado = df_completo['Monto_Cobrado'].sum()
                    df_errores = df_completo[df_completo['Estatus'] != '✅ Correcto']
                    monto_en_riesgo = df_errores['Monto_Cobrado'].sum()
                    
                    kpi1.metric("Total Operaciones Auditadas", f"{total_operaciones}")
                    kpi2.metric("Total Facturado (Cobrado)", f"${total_facturado:,.2f}")
                    kpi3.metric("Monto en Alerta/Riesgo", f"${monto_en_riesgo:,.2f}", delta_color="inverse", delta=f"{len(df_errores)} incidencias")
                    
                    gcol1, gcol2 = st.columns(2)
                    
                    with gcol1:
                        ventas_aduana = df_completo.groupby('Aduana')['Monto_Cobrado'].sum().reset_index()
                        fig1 = px.bar(ventas_aduana, x='Aduana', y='Monto_Cobrado', title="Venta Total por Aduana", color='Aduana', text_auto='.2s')
                        st.plotly_chart(fig1, use_container_width=True)
                        
                    with gcol2:
                        estatus_count = df_completo['Estatus'].value_counts().reset_index()
                        estatus_count.columns = ['Estatus', 'Cantidad']
                        fig2 = px.pie(estatus_count, names='Estatus', values='Cantidad', title="Distribución de Estatus", hole=0.4, color='Estatus', 
                                        color_discrete_map={'✅ Correcto':'#00CC96', '🚨 Error':'#EF553B', '⚠️ Alerta':'#FECB52', '🚨 Sobrecobro':'#FFA15A', '🚨 Subcobro':'#AB63FA'})
                        st.plotly_chart(fig2, use_container_width=True)
                        
                st.divider()
                st.download_button(
                    label="📥 Descargar Reporte de Facturación (Excel)",
                    data=excel_reporte,
                    file_name="Reporte_Auditoria_Facturacion.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
            except Exception as e:
                st.error(f"🚨 Ocurrió un error: {e}")
        else:
            st.warning("⚠️ Faltan archivos por subir.")

elif opcion_auditor == "Auditor Cruces EILO":
    st.title("🚚 Plataforma de auditoria Cruces EILO")
    st.markdown("Cruza el reporte operativo de facturas de clientes contra lo facturado por compras en EILO.")

    col1, col2 = st.columns(2)
    with col1:
        archivo_cruces = st.file_uploader("🧾 1. Facturas (Documentos por referencia EILO)", type=["xlsx", "xls", "csv"])
    with col2:
        archivo_facturas = st.file_uploader("🛒 2. FacturasCompras (Compras EILO)", type=["xlsx", "xls", "csv"])

    st.divider()

    if st.button("🚀 Ejecutar Auditoría de Cruces", use_container_width=True):
        if archivo_cruces and archivo_facturas:
            barra_progreso = st.progress(0)
            with st.spinner("⏳ Analizando e identificando referencias..."):
                try:
                    excel_cruces, df_resumen = auditar_cruces_eilo(archivo_cruces, archivo_facturas, barra_progreso)
                    
                    st.success("✅ ¡Auditoría de Cruces completada!")
                    
                    st.markdown("---")
                    st.header("📊 Resumen de Cruces")
                    
                    # Gráfica del resumen de EILO
                    if not df_resumen.empty:
                        fig = px.bar(df_resumen, x='Estatus de Auditoría', y='Cantidad de Eventos', 
                                     title="Distribución de Estatus de Cruces", color='Estatus de Auditoría', text_auto=True)
                        st.plotly_chart(fig, use_container_width=True)

                    st.divider()
                    st.download_button(
                        label="📥 Descargar Reporte de Cruces (Excel)",
                        data=excel_cruces,
                        file_name="Resultado_Conciliacion_Auditoria.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                except Exception as e:
                    st.error(f"🚨 {e}")
        else:
            st.warning("⚠️ Sube ambos reportes para continuar.")

else:
    st.title("🏢 Auditor ELI")
    st.info("🚧 Este módulo se encuentra actualmente en fase de diseño y desarrollo. Estará disponible en futuras actualizaciones.")